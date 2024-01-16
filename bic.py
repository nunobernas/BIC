import sys
import os
import openpyxl
import requests

from datetime import datetime
from openpyxl.styles import Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from copy import copy
from jira import JIRA
from urllib.request import urlopen

username = "xxxxxxx@xxxxx.xxx"
base_url = "https://xxxxxxxx.atlassian.net/"
api_key = "xxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxx"
project_name = 'xxxxxxxx'
project_key = "XXX"

headers = {"Content-Type": "application/json", }


def internet_connection():
    try:
        urlopen('https://www.google.com', timeout=1)

        return True

    except Exception as e:
        print("[SYSTEM]: Couldn't connect to the Internet.\n[INFO]: Check your "
              "internet connection!\n")
        print("[INFO]: Exception: " + str(e))


def check_columns(ws, found_valid, found_issueType):
    for cell in ws[1]:
        if cell.value == "Issue Type":
            found_issueType = 1

        if cell.value == "Valid ":
            found_valid = 1

        if found_valid and found_issueType:
            return True

    return False


def get_area_values(sheet):
    return [cell.value for row in sheet.iter_rows(min_col=1, max_col=1, min_row=2) for cell in row if
            cell.value is not None]


def create_columns(found_valid, found_issueType, size, ws):
    if found_valid == 0:
        for data_validation in ws.data_validations.dataValidation:
            ws.data_validations.dataValidation.remove(data_validation)

        ws.column_dimensions['G'].width = 25
        ws.column_dimensions['K'].width = 25

        dropdown_valid = ["Non-valid - to Report", "Non-valid - Not for "
                                                   "Reporting", "Valid", "Not Applicable"]

        ws['G1'].value = "Valid "

        dv = DataValidation(type="list", formula1=f'"{",".join(dropdown_valid)}"')
        ws.add_data_validation(dv)
        dv.add(f'G2:G{size}')

    if found_issueType == 0:
        dropdown_issueValues = ["Bug", "New Feature", "Task", "Improvement"]

        ws.insert_cols(8, amount=1)
        ws.column_dimensions['H'].width = 15

        ws['H1'].value = "Issue Type"
        ws["H1"].font = copy(ws["D1"].font)
        ws["H1"].border = copy(ws["D1"].border)
        ws["H1"].fill = copy(ws["D1"].fill)
        ws["H1"].number_format = copy(ws["D1"].number_format)
        ws["H1"].protection = copy(ws["D1"].protection)
        ws["H1"].alignment = copy(ws["D1"].alignment)

        dv = DataValidation(
            type="list", formula1=f'"{",".join(dropdown_issueValues)}"')

        ws.add_data_validation(dv)
        dv.add(f'H2:H{size}')

        ws[f'H{size}'].border = Border(
            bottom=Side(border_style="medium", color="000000"),
        )


def get_value_between(number):
    values = number.split(".")
    return values


def replace_areas(all_entries, total_sheet_areas):
    beforeValueOne = 1
    beforeValueTwo = 1
    count = 0
    sorted_areas = []

    for sheet_area in total_sheet_areas:
        for area in sheet_area:
            sorted_areas.append(area)

    for row in all_entries:
        if row[1] is None:
            break

        values = get_value_between(row[1])

        values[0] = int(values[0])
        values[1] = int(values[1])

        if values[1] > beforeValueOne:
            count += 1
        if values[0] > beforeValueTwo:
            count += 1

        beforeValueTwo = int(values[0])
        beforeValueOne = int(values[1])

        row[0] = sorted_areas[count]


def print_entries(valid_rows, not_applicable_rows, null_rows, to_report,
                  not_report):
    if len(valid_rows) == 0:
        print("\n\n[INFO]: There are no 'Valid' values!\n")
    else:
        print("\n\n[INFO]: Rows with 'Valid' values:\n")

        for row_values in valid_rows:
            print(row_values)

    if len(to_report) == 0:
        print("\n\n[INFO]: There are no 'Non-valid - to Report' "
              "values!\n")
    else:
        print("\n\n[INFO]: Rows with 'Non-valid - to Report' values\n")
        for row_values in to_report:
            print(row_values)

    if len(not_report) == 0:
        print("\n\n[INFO]: There are no 'Non-valid - not for Reporting' "
              "values!\n")
    else:
        print("\n\n[INFO]: Rows with 'Non-valid - not for Reporting' values\n")
        for row_values in not_report:
            print(row_values)

    if len(not_applicable_rows) == 0:
        print("\n\n[INFO]: There are no 'Not Applicable' values!\n")
    else:
        print("\n\n[INFO]: Rows with 'Not Applicable' values:\n")
        for row_values in not_applicable_rows:
            print(row_values)

    if len(null_rows) == 0:
        print("\n\n[INFO]: There are no 'Null' values!\n")
    else:
        print("\n\n[INFO]: Rows with 'Null' values: \n")
        for row_values in null_rows:
            print(row_values)


def print_statistics(valid_rows, non_valid_rows, not_applicable_rows, null_rows,
                     to_report):
    print("\n\n= = = = = Statistics (SECURITY CONTROLS) = = = = = ")
    print("\n\nValid: " + str(len(valid_rows)) + "\n")
    print("Non-valid - to Report: " + str(len(to_report)) + "\n")
    print("Non-valid (Total): " + str(len(non_valid_rows)) + "\n")
    print("Not Applicable: " + str(len(not_applicable_rows)) + "\n")
    print("Null: " + str(len(null_rows)))
    print("\n\n= = = = = = = = = = = = = = = = = = = = = = = = =\n\n")


def define_issues(to_report, sheet_names, filename, non_valids):
    issues_defined = []
    count = 1
    for iss in to_report:
        chapter = sheet_names[int(get_value_between(iss[1])[0]) - 1]

        if iss[6] == "Non-valid - to Report":
            print("= = = = = ISSUE [#" + str(count) + "] = = = = =")
            print("Chapter: " + chapter + ";\nArea: " + str(iss[0]) + ";  \nID: "
                  + str(iss[1]) + "; \nASVS Level: " + str(iss[2]) + ";  \nVerification" +
                  "Requirement: " + str(iss[5]) + ";  \nIssue Type: " + str(iss[7]) +
                  "\nReport: " + str(iss[9]) + "\n\n")

            issue_to_add = [chapter, str(iss[0]), str(iss[1]), str(iss[2]),
                            str(iss[5]), str(iss[7]), str(iss[9])]

            count += 1
            issues_defined.append(issue_to_add)

    if len(issues_defined) == 0:
        print("[SYSTEM]: There are no issues to report!")

        if len(non_valids) == 0:
            print("[SYSTEM]: Make sure you have them selected with "
                  "'Non-valid - to Report'.")

        else:
            print("[SYSTEM]: No more issues to add to the Jira board")
        print("\n[INFO]: No file was generated as there are no issues to report."
              "\n\n")

    else:
        count = 1
        with (open(filename, "w") as file):
            for iss in issues_defined:
                op_text = "\n\nIssue [# " + str(count) + "]: " + str(iss) + "\n\n"
                file.write(op_text)
                count += 1

    return issues_defined


def get_jira_issues():
    issuesInJira = []

    jiraOptions = {'server': base_url}

    jira = JIRA(options=jiraOptions, basic_auth=(username, api_key))

    for iss in jira.search_issues(jql_str='project = ' + project_name):
        issuesInJira.append(iss.fields.summary)

    return issuesInJira


def create_jira_issue(issue_summary, issue_description, issueType, prio):
    issue_data = {
        "fields": {
            "project": {"key": project_key},
            "summary": issue_summary,
            "description": issue_description,
            "issuetype": {
                "name": issueType
            },
        }
    }

    response = requests.post(
        f"{base_url}/rest/api/2/issue",
        headers=headers,
        json=issue_data,
        auth=(username, api_key)
    )

    if response.status_code == 201:
        issue_to_send = response.json()
        issue_key = issue_to_send["key"]
        print(f"\nIssue created: {issue_key}")
        url = base_url + "jira/core/projects/" + project_key + "/board/" + issue_key
        print(f"Issue Link: {url}")

        update_url = f"{base_url}/rest/api/2/issue/{issue_key}"
        auth = (username, api_key)

        priority_data = {
            "update": {"priority": [{"set": {"name": prio}}]}
        }

        response = requests.put(update_url, headers=headers, json=priority_data, auth=auth)

        if response.status_code == 204:
            print(f"Priority updated successfully for issue {issue_key}")
        else:
            print(
                f"Failed to update priority. Status code: {response.status_code}, Response: {response.text}"
            )
    else:
        print(
            f"Failed to create JIRA issue. Status code: {response.status_code}, Error: {response.text}"
        )


def main():
    if len(sys.argv) < 3:
        print("[SYSTEM]: Invalid Syntax!\n[SYSTEM]: Use - python tool.py "
              "<excel_path> <project_name>")
        sys.exit(1)

    excel_name = '/' + sys.argv[1]
    parts = excel_name.split('/')
    excel_filename = parts[-1]
    excel_path = os.getcwd() + excel_name
    now = datetime.now()
    dt_string = now.strftime("%d-%m-%Y__%Hh%Mm%Ss")
    filename = './Reports/' + sys.argv[2] + '-' + dt_string + '.txt'

    non_valid_rows = []
    to_report = []
    not_report = []
    valid_rows = []
    not_applicable_rows = []
    null_rows = []
    all_entries = []

    total_sheet_areas = []

    number_of_rows = 1

    found_valid = 0
    found_issueType = 0

    control = 1

    issues_to_jira = None

    try:
        wb = openpyxl.load_workbook(excel_path)

        sheet_names = wb.sheetnames

        sheet_names.remove("ASVS Results")

        if "Export Summary" in sheet_names:
            sheet_names.remove("Export Summary")

        for sheet_name in sheet_names:

            found_issueType = 0
            found_valid = 0

            ws = wb[sheet_name]

            if check_columns(ws, found_valid, found_issueType):
                control = 1
                for row in ws.iter_rows(min_col=1, max_col=10, min_row=2):
                    row_values = [cell.value for cell in row]
                    value = row_values[6]

                    if row_values[1] is None:
                        break

                    all_entries.append(row_values)

                    if value == "Valid":
                        valid_rows.append(row_values)
                    elif value == "Not Applicable":
                        not_applicable_rows.append(row_values)
                    elif value is None:
                        null_rows.append(row_values)
                    else:
                        non_valid_rows.append(row_values)
                        if value == "Non-valid - Not for Reporting":
                            not_report.append(row_values)
                        if value == "Non-valid - to Report":
                            to_report.append(row_values)

                column_a_values = get_area_values(ws)
                total_sheet_areas.append(column_a_values)

            else:

                for row in ws.iter_rows(min_row=2, max_col=11, min_col=7,
                                        max_row=ws.max_row):
                    for cell in row:
                        if cell.coordinate in ws.merged_cells:
                            ws.unmerge_cells(cell.coordinate)
                        cell.value = None

                create_columns(found_valid, found_issueType, ws.max_row, ws)
                control = 0

        if control == 0:
            print("\n[SYSTEM]: Excel structure has been updated!")
            print("[SYSTEM]: You now can fill the Excel '" + str(
                excel_filename) + "' and start to generate issues to your Project\n")
            print("[INFO]: No file was generated as there are no issues to report at this stage\n\n")

        else:
            if not os.path.exists("./Reports") and len(non_valid_rows) != 0:
                os.makedirs("./Reports")

            replace_areas(all_entries, total_sheet_areas)

            # print_entries(valid_rows, not_applicable_rows, null_rows, to_report, not_report)
            print_statistics(valid_rows, non_valid_rows, not_applicable_rows, null_rows, to_report)

            if (to_report is not None) and (internet_connection() is True):
                summaryInJira = get_jira_issues()

                filtered_issues = []

                for i in to_report:
                    if "OWASP-ASVS #" + str(i[1]) not in summaryInJira:
                        filtered_issues.append(i)

                issues_to_send = filtered_issues

                issues_to_jira = define_issues(issues_to_send, sheet_names, filename, non_valid_rows)

        wb.save(excel_path)

        if 'wb' in locals():
            wb.close()

        return issues_to_jira

    except Exception as e:
        print(f"[SYSTEM]: {e}\n")


if __name__ == "__main__":

    issues_to_report = main()

    if (issues_to_report is not None) and (internet_connection() is True):
        for issue in issues_to_report:
            if issue[5] == "None" or issue[5] is None:
                issue[5] = "Task"
            if issue[5] == "Bug":
                priority = "Highest"
            elif issue[5] == "New Feature":
                priority = "High"
            elif issue[5] == "Task":
                priority = "Medium"
            else:
                priority = "Low"

            summary = "OWASP-ASVS #" + str(issue[2])
            description = issue[6] + "\n\n\nChapter: " + str(issue[0]) + "\nArea: " + str(
                issue[1]) + "\nRequirement Description: " + issue[4]

            create_jira_issue(summary, description, issue[5], priority)
